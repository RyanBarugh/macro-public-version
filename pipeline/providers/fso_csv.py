from __future__ import annotations

import base64
import io
import time
from typing import Any, Optional, Tuple

import pandas as pd
import requests

from .base import BaseProvider
from ..engine.series import SeriesDef
from ..engine.logger import get_logger

logger = get_logger(__name__)

MAX_DROP_PCT = 5.0

# Swiss Federal Statistical Office — DAM (Digital Asset Management) API
# Direct CSV downloads from dam-api.bfs.admin.ch
# No authentication required.
BASE_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets"


class FsoCsvProvider(BaseProvider):
    """
    Swiss Federal Statistical Office (BFS) DAM API provider.

    Downloads CSV or XLSX files from the BFS Digital Asset Management system.
    Each dataset has a fixed asset ID; the URL pattern is:
        GET {BASE_URL}/{asset_id}/master

    CSV mode (default):
        asset_id     : int or str — the DAM asset number (e.g. 36210011)
        The CSV is expected to have a PERIOD column in YYYY-Qn or YYYY-MM format.

    XLSX mode (format: "xlsx"):
        asset_id     : int or str — the DAM asset number
        format       : "xlsx"
        sheet        : str — worksheet name (e.g. "Indices")
        row_code     : str — sector code in column A to select (e.g. "B-E")
        Wide format: quarters as column headers (YYYY/I .. YYYY/IV), sectors as rows.
    """

    def fetch(
        self,
        series_def: SeriesDef,
        session: requests.Session,
        timeout: Tuple[float, float],
        api_key: Optional[str],  # unused — BFS is keyless
        start: str,
    ) -> Any:
        meta = series_def.meta or {}
        asset_id = meta.get("asset_id")
        fmt = meta.get("format", "csv")

        if not asset_id:
            raise RuntimeError(
                f"FSO CSV series_id={series_def.series_id} missing "
                f"'asset_id' in meta"
            )

        url = f"{BASE_URL}/{asset_id}/master"

        logger.info(
            "Fetching provider=fso_csv series_id=%s asset_id=%s format=%s",
            series_def.series_id, asset_id, fmt,
        )

        start_ts = time.time()
        try:
            response = session.get(url, timeout=timeout)
            elapsed_ms = int((time.time() - start_ts) * 1000)
            logger.info(
                "HTTP provider=fso_csv series_id=%s status=%s elapsed_ms=%d",
                series_def.series_id, response.status_code, elapsed_ms,
            )
            response.raise_for_status()

            if fmt == "xlsx":
                raw_bytes = response.content
                if not raw_bytes or len(raw_bytes) < 100:
                    raise RuntimeError(
                        f"FSO XLSX returned empty/tiny response for "
                        f"series_id={series_def.series_id}"
                    )
                logger.info(
                    "Fetched provider=fso_csv series_id=%s bytes=%d",
                    series_def.series_id, len(raw_bytes),
                )
                return {
                    "format": "xlsx",
                    "bytes_b64": base64.b64encode(raw_bytes).decode("ascii"),
                    "sheet": meta.get("sheet", "Indices"),
                    "row_code": meta.get("row_code"),
                }

            # CSV path (original)
            text = response.content.decode("utf-8-sig")
            if not text or len(text.strip()) == 0:
                raise RuntimeError(
                    f"FSO CSV returned empty response for "
                    f"series_id={series_def.series_id}"
                )

            row_count = max(text.count("\n") - 1, 0)
            logger.info(
                "Fetched provider=fso_csv series_id=%s rows=%d",
                series_def.series_id, row_count,
            )
            return text

        except Exception:
            logger.exception(
                "Failed provider=fso_csv series_id=%s", series_def.series_id
            )
            raise

    def clean(
        self,
        raw_payload: Any,
        series_id: str,
        strict: bool = True,
    ) -> pd.DataFrame:
        # Dispatch: XLSX returns a dict from fetch, CSV returns a string
        if isinstance(raw_payload, dict) and raw_payload.get("format") == "xlsx":
            return self._clean_xlsx(raw_payload, series_id, strict)
        return self._clean_csv(raw_payload, series_id, strict)

    # ── XLSX path (wide-format quarterly indices) ──────────────────

    def _clean_xlsx(
        self,
        payload: dict,
        series_id: str,
        strict: bool = True,
    ) -> pd.DataFrame:
        """
        Parse FSO wide-format XLSX.

        Supports two header layouts:
          1. Combined: single row with "YYYY/I", "YYYY/II", ... (INDPAU)
          2. Split: year row (1991, 1992, ...) + quarter row (I, II, III, IV)
             with merged year cells spanning 4 columns (BESTA)

        row_code selects which sector row to extract (col A match).
        """
        import openpyxl

        sheet_name = payload.get("sheet", "Indices")
        row_code = payload.get("row_code")

        if not row_code:
            raise ValueError(
                f"FSO XLSX series_id={series_id} missing 'row_code'"
            )

        logger.info(
            "Cleaning provider=fso_csv(xlsx) series_id=%s sheet=%s row_code=%s",
            series_id, sheet_name, row_code,
        )

        # Use data_only=True but NOT read_only — read_only breaks merged cells
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(payload["bytes_b64"])),
            data_only=True,
        )
        ws = wb[sheet_name]

        # ── Detect header format ──────────────────────────────────
        # Scan first 10 rows for either combined "YYYY/Q" or split year+quarter
        col_timestamps = {}
        data_start_row = None
        _q_map = {"I": "01", "II": "04", "III": "07", "IV": "10"}

        # Try combined format first: "YYYY/I" or "I-YYYY" style headers
        for row_idx in range(1, min(11, ws.max_row + 1)):
            cells = [c.value for c in ws[row_idx]]
            sample = [
                c for c in cells[1:] if c and isinstance(c, str)
                and ("/" in c or "-" in c)
            ]
            if len(sample) >= 4:
                # Determine separator and order
                s0 = sample[0]
                if "/" in s0:
                    sep = "/"
                    # "YYYY/I" → year first
                    order = "year_first"
                elif "-" in s0:
                    sep = "-"
                    # Check if "I-1975" (quarter first) or "1975-I" (year first)
                    left = s0.split("-")[0].strip()
                    order = "quarter_first" if left in _q_map else "year_first"
                else:
                    continue

                # Verify at least 4 match the expected pattern
                matched = 0
                for c in sample:
                    parts = c.split(sep)
                    if len(parts) == 2:
                        if order == "year_first":
                            q_part = parts[1].strip()
                        else:
                            q_part = parts[0].strip()
                        if q_part in _q_map:
                            matched += 1
                if matched < 4:
                    continue

                logger.info(
                    "Detected combined header at row %d (sep='%s', order=%s)",
                    row_idx, sep, order,
                )
                for i in range(1, len(cells)):
                    h = cells[i]
                    if not h or not isinstance(h, str) or sep not in h:
                        continue
                    try:
                        parts = h.split(sep)
                        if order == "year_first":
                            year = parts[0].strip()
                            q_roman = parts[1].strip()
                        else:
                            q_roman = parts[0].strip()
                            year = parts[1].strip()
                        month = _q_map.get(q_roman)
                        if month:
                            col_timestamps[i] = pd.Timestamp(f"{year}-{month}-01")
                    except Exception:
                        continue
                data_start_row = row_idx + 1
                break

        # Try split format: year row + quarter row
        if not col_timestamps:
            year_row_idx = None
            q_row_idx = None
            for row_idx in range(1, min(11, ws.max_row + 1)):
                cells = [c.value for c in ws[row_idx]]
                # Year row: has integers >= 1900 in data columns
                years_found = [c for c in cells[3:] if isinstance(c, (int, float)) and 1900 <= c <= 2100]
                if len(years_found) >= 3 and year_row_idx is None:
                    year_row_idx = row_idx
                    continue
                # Quarter row: has roman numerals
                romans_found = [c for c in cells[3:] if c in ("I", "II", "III", "IV")]
                if len(romans_found) >= 4 and q_row_idx is None:
                    q_row_idx = row_idx
                    break

            if year_row_idx and q_row_idx:
                logger.info(
                    "Detected split year/quarter headers at rows %d/%d",
                    year_row_idx, q_row_idx,
                )
                year_cells = [c.value for c in ws[year_row_idx]]
                q_cells = [c.value for c in ws[q_row_idx]]

                # Forward-fill years across columns
                current_year = None
                for i in range(3, len(year_cells)):
                    y = year_cells[i]
                    if isinstance(y, (int, float)) and 1900 <= y <= 2100:
                        current_year = int(y)
                    q_label = q_cells[i] if i < len(q_cells) else None
                    if current_year and q_label in _q_map:
                        month = _q_map[q_label]
                        col_timestamps[i] = pd.Timestamp(f"{current_year}-{month}-01")

                data_start_row = q_row_idx + 1

        if not col_timestamps:
            wb.close()
            raise ValueError(
                f"FSO XLSX series_id={series_id}: could not detect "
                f"header format in sheet '{sheet_name}'"
            )

        # ── Find the target data row ──────────────────────────────
        values = None
        for row in ws.iter_rows(min_row=data_start_row, values_only=False):
            cell_vals = [c.value for c in row]
            if cell_vals[0] == row_code:
                values = cell_vals
                break

        wb.close()

        if values is None:
            raise ValueError(
                f"FSO XLSX series_id={series_id}: row_code='{row_code}' "
                f"not found in sheet '{sheet_name}'"
            )

        # ── Build time series ─────────────────────────────────────
        records = []
        for col_idx, ts in sorted(col_timestamps.items()):
            v = values[col_idx] if col_idx < len(values) else None
            if v is None or not isinstance(v, (int, float)):
                continue
            records.append({"time": ts, "value": round(float(v), 6)})

        if not records:
            raise ValueError(
                f"FSO XLSX series_id={series_id}: no valid data points "
                f"extracted for row_code='{row_code}'"
            )

        out = pd.DataFrame(records)
        out.insert(0, "series_id", series_id)
        out = out.sort_values("time").reset_index(drop=True)

        before = len(out)
        out = out.dropna(subset=["time", "value"])
        dropped = before - len(out)

        if dropped:
            drop_pct = (dropped / before) * 100
            logger.warning(
                "Dropped %d/%d rows (%.1f%%) for series_id=%s",
                dropped, before, drop_pct, series_id,
            )
            if strict and drop_pct > MAX_DROP_PCT:
                raise ValueError(
                    f"series_id={series_id}: dropped {drop_pct:.1f}% of rows "
                    f"(threshold={MAX_DROP_PCT}%). Possible format change."
                )

        logger.info(
            "Cleaned rows series_id=%s rows_out=%d range=%s to %s",
            series_id, len(out),
            out["time"].iloc[0].date(), out["time"].iloc[-1].date(),
        )
        out["time"] = out["time"].dt.strftime("%Y-%m-%d")
        return out

    # ── CSV path (original long-format) ────────────────────────────

    def _clean_csv(
        self,
        raw_payload: Any,
        series_id: str,
        strict: bool = True,
    ) -> pd.DataFrame:
        """
        Parse FSO CSV response.

        Expected format (Swiss Wage Index example):
            "PERIOD","WAGE_TYPE","VALUE_P","STATUS_VALUE_P"
            "1991-Q1","N","7.8","A"
            ...
        """
        logger.info(
            "Cleaning provider=fso_csv series_id=%s strict=%s",
            series_id, strict,
        )

        if not raw_payload:
            raise ValueError(
                f"FSO CSV payload empty for series_id={series_id}"
            )

        # Strip UTF-8 BOM if present (response.text may retain it as \ufeff)
        text = raw_payload.lstrip("\ufeff")

        df = pd.read_csv(io.StringIO(text))

        logger.info(
            "Raw rows series_id=%s rows_in=%d columns=%s",
            series_id, len(df), list(df.columns),
        )

        # Normalise column names to uppercase stripped
        df.columns = [c.strip().strip('"').upper() for c in df.columns]

        if "PERIOD" not in df.columns:
            raise ValueError(
                f"FSO CSV missing PERIOD column for series_id={series_id}. "
                f"Columns: {list(df.columns)}"
            )

        # Apply optional filter (e.g. WAGE_TYPE == "N")
        # Read filter from the series meta via series_id lookup — but we only
        # have the raw payload and series_id here. Use a convention:
        # the caller passes filter info via the value_col name.
        # For simplicity, detect known filter columns automatically.
        filter_cols = [
            ("WAGE_TYPE", None),
            ("INDICATOR", None),
        ]
        for col, _ in filter_cols:
            if col in df.columns:
                unique_vals = df[col].unique()
                if len(unique_vals) > 1:
                    logger.info(
                        "Multiple %s values found: %s — using first: %s",
                        col, unique_vals.tolist(), unique_vals[0],
                    )
                    df = df[df[col] == unique_vals[0]]

        # Find the value column — try common names
        value_col = None
        for candidate in ["VALUE_P", "VALUE", "DATA_VALUE", "OBS_VALUE"]:
            if candidate in df.columns:
                value_col = candidate
                break

        if value_col is None:
            raise ValueError(
                f"FSO CSV CSV: no value column found for series_id={series_id}. "
                f"Columns: {list(df.columns)}"
            )

        # Parse PERIOD — handle YYYY-Qn and YYYY-MM
        sample = str(df["PERIOD"].dropna().iloc[0]) if not df.empty else ""

        if "-Q" in sample:
            _q_map = {"Q1": "01", "Q2": "04", "Q3": "07", "Q4": "10"}

            def _parse_quarter(val: str) -> pd.Timestamp:
                try:
                    year, q = str(val).strip().split("-")
                    month = _q_map.get(q.upper())
                    if not month:
                        return pd.NaT
                    return pd.Timestamp(f"{year}-{month}-01")
                except Exception:
                    return pd.NaT

            df["time"] = df["PERIOD"].apply(_parse_quarter)
            logger.info(
                "Parsed quarterly PERIOD format for series_id=%s", series_id
            )
        else:
            df["time"] = pd.to_datetime(
                df["PERIOD"], format="%Y-%m", errors="coerce"
            )

        df["value"] = pd.to_numeric(df[value_col], errors="coerce")

        out = df[["time", "value"]].copy()
        out.insert(0, "series_id", series_id)

        before = len(out)
        out = out.dropna(subset=["time", "value"]).sort_values("time")
        dropped = before - len(out)

        if dropped:
            drop_pct = (dropped / before) * 100
            logger.warning(
                "Dropped %d/%d rows (%.1f%%) for series_id=%s",
                dropped, before, drop_pct, series_id,
            )
            if strict and drop_pct > MAX_DROP_PCT:
                raise ValueError(
                    f"series_id={series_id}: dropped {drop_pct:.1f}% of rows "
                    f"(threshold={MAX_DROP_PCT}%). Possible format change."
                )

        dupes = out[out["time"].duplicated(keep=False)]
        if not dupes.empty:
            dupe_times = dupes["time"].unique().tolist()
            raise ValueError(
                f"series_id={series_id}: duplicate time values: {dupe_times}"
            )

        logger.info(
            "Cleaned rows series_id=%s rows_out=%d", series_id, len(out)
        )
        out["time"] = out["time"].dt.strftime("%Y-%m-%d")
        return out