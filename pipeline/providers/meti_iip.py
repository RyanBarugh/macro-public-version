from __future__ import annotations

import io
import re
import time
from typing import Any, Optional, Tuple

import openpyxl
import pandas as pd
import requests

from .base import BaseProvider
from ..engine.series import SeriesDef
from ..engine.logger import get_logger

logger = get_logger(__name__)

# METI Indices of Industrial Production (IIP), base 2020
# Seasonally Adjusted Index, By industry, Monthly
# Static URL — updated in place each month
# Sheet: Production
# Row 2 (0-indexed): date headers as YYYYMM integers or 'p YYYYMM' strings (p = preliminary)
# Col 0: Item_Number, Col 1: Item_Name, Col 2: Weight, Col 3+: monthly values

FILE_URL = (
    "https://www.meti.go.jp/english/statistics/tyo/iip/xls/b2020_gsm1e.xlsx"
)
SHEET_NAME = "Production"
DATE_ROW_IDX   = 2   # row containing YYYYMM date headers
DATA_START_COL = 3   # first column with index values (after Item_Number, Item_Name, Weight)
ITEM_COL       = 0   # column containing Item_Number

MAX_DROP_PCT = 5.0


class MetiIipProvider(BaseProvider):
    """
    METI Indices of Industrial Production — SA index by industry (2020=100).

    Fetches b2020_gsm1e.xlsx (static URL, updated monthly in place) and
    extracts the row matching `item_number` from the Production sheet.

    Date headers are YYYYMM integers for confirmed months, or 'p YYYYMM'
    strings for the latest preliminary month. Both are handled.

    series.json fields:
        item_number   : int   — 1000000000 (total) or 1100000000 (manufacturing)
        db_series_id  : str   — series_id written to DB
        insert_only   : true  — preliminary source, never overwrites existing rows
        backfill.start: "YYYY-MM"

    No API key required.
    """

    def fetch(
        self,
        series_def: SeriesDef,
        session: requests.Session,
        timeout: Tuple[float, float],
        api_key: Optional[str],
        start: str,
    ) -> Any:
        series_id = series_def.series_id
        meta = series_def.meta or {}

        item_number = meta.get("item_number")
        if item_number is None:
            raise RuntimeError(
                f"provider=meti_iip series_id={series_id}: missing 'item_number' in series meta"
            )
        item_number = int(item_number)

        logger.info(
            "Fetching provider=meti_iip series_id=%s item_number=%d url=%s",
            series_id, item_number, FILE_URL,
        )

        t0 = time.time()
        try:
            resp = session.get(FILE_URL, timeout=timeout)
            elapsed_ms = int((time.time() - t0) * 1000)
            logger.info(
                "HTTP provider=meti_iip series_id=%s status=%s elapsed_ms=%d",
                series_id, resp.status_code, elapsed_ms,
            )
            resp.raise_for_status()
        except Exception:
            logger.exception(
                "provider=meti_iip failed url=%s series_id=%s", FILE_URL, series_id
            )
            raise

        records = _parse_iip(resp.content, item_number, series_id)

        logger.info(
            "provider=meti_iip series_id=%s item_number=%d raw_rows=%d",
            series_id, item_number, len(records),
        )

        return {"records": records, "start_filter": start}

    def clean(
        self,
        raw_payload: Any,
        series_id: str,
        strict: bool = True,
    ) -> pd.DataFrame:
        logger.info(
            "Cleaning provider=meti_iip series_id=%s strict=%s", series_id, strict
        )

        records = raw_payload.get("records", [])
        start_filter = raw_payload.get("start_filter")

        if not records:
            raise ValueError(
                f"provider=meti_iip series_id={series_id}: empty records"
            )

        df = pd.DataFrame(records)
        df["time"]  = pd.to_datetime(df["time"],  errors="coerce")
        df["value"] = pd.to_numeric(df["value"], errors="coerce")

        before = len(df)
        df = df.dropna(subset=["time", "value"])
        dropped = before - len(df)
        if dropped:
            drop_pct = (dropped / before) * 100
            logger.warning(
                "provider=meti_iip series_id=%s dropped %d/%d rows (%.1f%%)",
                series_id, dropped, before, drop_pct,
            )
            if strict and drop_pct > MAX_DROP_PCT:
                raise ValueError(
                    f"provider=meti_iip series_id={series_id}: dropped {drop_pct:.1f}% "
                    f"of rows (threshold={MAX_DROP_PCT}%). Possible format change."
                )

        if start_filter:
            cutoff = pd.to_datetime(f"{start_filter}-01")
            df = df[df["time"] >= cutoff]

        df = df.sort_values("time").drop_duplicates(subset=["time"], keep="last")

        dupes = df[df["time"].duplicated(keep=False)]
        if not dupes.empty:
            raise ValueError(
                f"provider=meti_iip series_id={series_id}: duplicate times: "
                f"{dupes['time'].unique().tolist()}"
            )

        out = df[["time", "value"]].copy()
        out.insert(0, "series_id", series_id)
        out["time"] = out["time"].dt.strftime("%Y-%m-%d")

        logger.info(
            "provider=meti_iip series_id=%s rows_out=%d", series_id, len(out)
        )
        return out


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _parse_iip(content: bytes, item_number: int, series_id: str) -> list[dict]:
    """
    Parse Production sheet from b2020_gsm1e.xlsx bytes.

    1. Parse date headers from DATE_ROW_IDX — int or 'p YYYYMM' string.
    2. Find data row where col[ITEM_COL] == item_number.
    3. Zip dates with values and return records.

    Raises clearly if item_number not found — canary for layout changes.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as e:
        raise RuntimeError(
            f"provider=meti_iip series_id={series_id}: failed to open workbook: {e}"
        ) from e

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"provider=meti_iip series_id={series_id}: sheet '{SHEET_NAME}' not found. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Parse date headers
    date_row = rows[DATE_ROW_IDX]
    dates: list[tuple[int, str]] = []
    for col_idx, cell in enumerate(date_row):
        if col_idx < DATA_START_COL:
            continue
        time_str = _parse_date_header(cell)
        if time_str:
            dates.append((col_idx, time_str))

    if not dates:
        raise RuntimeError(
            f"provider=meti_iip series_id={series_id}: no date headers found "
            f"in row {DATE_ROW_IDX}. Layout may have changed."
        )

    logger.info(
        "provider=meti_iip series_id=%s dates=%s to %s months=%d",
        series_id, dates[0][1], dates[-1][1], len(dates),
    )

    # Find data row by item_number
    data_row = None
    for row in rows[DATE_ROW_IDX + 1:]:
        if row and row[ITEM_COL] == item_number:
            data_row = row
            break

    if data_row is None:
        raise RuntimeError(
            f"provider=meti_iip series_id={series_id}: item_number={item_number} "
            f"not found in sheet '{SHEET_NAME}'. Layout may have changed."
        )

    logger.info(
        "provider=meti_iip series_id=%s resolved item_number=%d name='%s'",
        series_id, item_number, data_row[1],
    )

    # Extract records
    records = []
    for col_idx, time_str in dates:
        if col_idx >= len(data_row):
            continue
        raw_val = data_row[col_idx]
        if raw_val is None:
            continue
        try:
            value = float(raw_val)
        except (TypeError, ValueError):
            logger.warning(
                "provider=meti_iip series_id=%s unparseable value=%r at time=%s",
                series_id, raw_val, time_str,
            )
            continue
        records.append({"time": time_str, "value": value})

    return records


def _parse_date_header(cell: Any) -> str | None:
    """
    Parse a date header cell to "YYYY-MM-01".

    int/float : 201801       → "2018-01-01"
    str       : "p 202601"   → "2026-01-01"  (preliminary)
    str       : "202601"     → "2026-01-01"
    """
    if cell is None:
        return None

    if isinstance(cell, (int, float)):
        s = str(int(cell))
        if len(s) == 6 and s.isdigit():
            return f"{s[:4]}-{s[4:6]}-01"
        return None

    if isinstance(cell, str):
        s = re.sub(r'^p\s*', '', cell.strip())
        if len(s) == 6 and s.isdigit():
            return f"{s[:4]}-{s[4:6]}-01"

    return None